"""
Handler para generar reportes en Excel por comando /reportes
"""

import os
import json
import logging
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import io

from telebot import types
from utils.logger import get_logger
from config.database import db_manager
from core.os_detector import OSDetector

logger = get_logger(__name__)


class ReportGenerator:
    """Generador de reportes en Excel"""

    def __init__(self):
        self.reports_dir = "reports"
        self._ensure_reports_dir()

    def _ensure_reports_dir(self):
        """Asegura que el directorio de reports exista"""
        os.makedirs(self.reports_dir, exist_ok=True)

    def generate_usage_report(self, user_id: int, username: str) -> str:
        """Genera reporte completo de uso en Excel"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"reporte_uso_{timestamp}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)

            # Crear workbook
            wb = Workbook()

            # Hoja 1: Resumen general
            ws1 = wb.active
            ws1.title = "Resumen General"

            # Encabezados
            headers = ["Fecha Generación", "Generado por", "Usuario ID", "Total Conexiones", "Tiendas Activas"]
            ws1.append(headers)

            # Obtener datos
            connection_stats = self._get_connection_stats()
            store_stats = self._get_store_stats()

            # Datos
            ws1.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                username,
                user_id,
                connection_stats.get('total_connections', 0),
                store_stats.get('active_stores', 0)
            ])

            # Hoja 2: Estadísticas de tiendas
            ws2 = wb.create_sheet("Estadísticas Tiendas")
            ws2.append(["Tienda", "Sistema", "Última Conexión", "Estado", "Conexiones Totales"])

            for store in store_stats.get('stores', []):
                ws2.append([
                    store.get('store_code', ''),
                    store.get('os_type', ''),
                    store.get('last_connection', ''),
                    store.get('status', ''),
                    store.get('connection_count', 0)
                ])

            # Hoja 3: Logs de uso
            ws3 = wb.create_sheet("Logs de Uso")
            ws3.append(["Fecha/Hora", "Usuario", "Acción", "Tienda", "Detalles"])

            logs = self._get_usage_logs()
            for log in logs[-100:]:  # Últimos 100 registros
                ws3.append([
                    log.get('timestamp', ''),
                    log.get('username', ''),
                    log.get('action', ''),
                    log.get('store', ''),
                    log.get('details', '')
                ])

            # Hoja 4: Gráficos
            ws4 = wb.create_sheet("Gráficos")

            # Crear gráficos
            chart_data = self._prepare_chart_data(store_stats)
            chart_image = self._create_charts(chart_data)

            if chart_image:
                img = ExcelImage(chart_image)
                ws4.add_image(img, 'A1')

            # Ajustar anchos de columna
            for ws in [ws1, ws2, ws3]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar archivo
            wb.save(filepath)

            logger.info(f"Reporte generado: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error generando reporte: {e}")
            return None

    def _get_connection_stats(self) -> Dict[str, Any]:
        """Obtiene estadísticas de conexiones"""
        try:
            # Leer logs de conexión
            log_file = "logs/bot.log"
            if not os.path.exists(log_file):
                return {"total_connections": 0}

            with open(log_file, 'r', encoding='utf-8') as f:
                content = f.read()

            # Contar conexiones
            connections = content.count("Conexión establecida") + content.count("conectado")

            return {
                "total_connections": connections,
                "log_file": log_file
            }

        except Exception as e:
            logger.error(f"Error obteniendo estadísticas: {e}")
            return {"total_connections": 0}

    def _get_store_stats(self) -> Dict[str, Any]:
        """Obtiene estadísticas de tiendas"""
        try:
            # Esta función necesitaría acceder a la base de datos de logs
            # Por ahora devolvemos datos de ejemplo
            stores = []

            # Ejemplo de tiendas comunes
            for i in range(1, 11):
                store_code = f"K{str(i).zfill(3)}"
                os_type = "Linux" if i % 2 == 0 else "Windows"

                stores.append({
                    "store_code": store_code,
                    "os_type": os_type,
                    "last_connection": (datetime.now() - timedelta(hours=i)).strftime("%Y-%m-%d %H:%M"),
                    "status": "Activa" if i < 8 else "Inactiva",
                    "connection_count": i * 3
                })

            return {
                "active_stores": len([s for s in stores if s["status"] == "Activa"]),
                "total_stores": len(stores),
                "stores": stores
            }

        except Exception as e:
            logger.error(f"Error obteniendo stats de tiendas: {e}")
            return {"active_stores": 0, "total_stores": 0, "stores": []}

    def _get_usage_logs(self) -> List[Dict[str, Any]]:
        """Obtiene logs de uso"""
        try:
            usage_file = "logs/bot_usage.log"
            if not os.path.exists(usage_file):
                return []

            logs = []
            with open(usage_file, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip():
                        try:
                            # Parsear línea de log: [timestamp] USER:123 NAME:usuario ACTION:accion DETAILS:detalles
                            parts = line.strip().split('] ')
                            if len(parts) >= 2:
                                timestamp = parts[0][1:]  # Quitar el primer '['
                                details = parts[1]

                                # Parsear detalles
                                user_id = ""
                                username = ""
                                action = ""
                                store = ""
                                details_text = ""

                                for item in details.split(' '):
                                    if item.startswith('USER:'):
                                        user_id = item[5:]
                                    elif item.startswith('NAME:'):
                                        username = item[5:]
                                    elif item.startswith('ACTION:'):
                                        action = item[7:]
                                    elif item.startswith('STORE:'):
                                        store = item[6:]
                                    elif item.startswith('DETAILS:'):
                                        details_text = item[8:]

                                logs.append({
                                    "timestamp": timestamp,
                                    "user_id": user_id,
                                    "username": username,
                                    "action": action,
                                    "store": store,
                                    "details": details_text
                                })
                        except:
                            continue

            return logs

        except Exception as e:
            logger.error(f"Error leyendo logs de uso: {e}")
            return []

    def _prepare_chart_data(self, store_stats: Dict) -> Dict:
        """Prepara datos para gráficos"""
        try:
            stores = store_stats.get('stores', [])

            # Distribución por sistema operativo
            os_counts = {}
            for store in stores:
                os_type = store.get('os_type', 'Desconocido')
                os_counts[os_type] = os_counts.get(os_type, 0) + 1

            # Estados de tiendas
            status_counts = {}
            for store in stores:
                status = store.get('status', 'Desconocido')
                status_counts[status] = status_counts.get(status, 0) + 1

            return {
                "os_distribution": os_counts,
                "status_distribution": status_counts,
                "stores": stores
            }

        except Exception as e:
            logger.error(f"Error preparando datos de gráficos: {e}")
            return {}

    def _create_charts(self, chart_data: Dict) -> io.BytesIO:
        """Crea gráficos y los guarda en BytesIO"""
        try:
            # Crear figura con múltiples subgráficos
            fig, axes = plt.subplots(2, 2, figsize=(12, 10))
            fig.suptitle('Reporte de Uso del Sistema KFC', fontsize=16)

            # Gráfico 1: Distribución de sistemas operativos
            ax1 = axes[0, 0]
            os_data = chart_data.get("os_distribution", {})
            if os_data:
                labels = list(os_data.keys())
                sizes = list(os_data.values())
                colors = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99']
                ax1.pie(sizes, labels=labels, colors=colors[:len(labels)], autopct='%1.1f%%')
                ax1.set_title('Distribución por Sistema Operativo')

            # Gráfico 2: Estados de tiendas
            ax2 = axes[0, 1]
            status_data = chart_data.get("status_distribution", {})
            if status_data:
                labels = list(status_data.keys())
                values = list(status_data.values())
                bars = ax2.bar(labels, values, color=['#4CAF50', '#F44336', '#FFC107'])
                ax2.set_title('Estados de Tiendas')
                ax2.set_ylabel('Cantidad')

                # Agregar valores encima de las barras
                for bar in bars:
                    height = bar.get_height()
                    ax2.text(bar.get_x() + bar.get_width() / 2., height,
                             f'{int(height)}', ha='center', va='bottom')

            # Gráfico 3: Top 5 tiendas más activas
            ax3 = axes[1, 0]
            stores = chart_data.get("stores", [])
            if stores:
                # Ordenar por conexiones
                stores_sorted = sorted(stores, key=lambda x: x.get('connection_count', 0), reverse=True)[:5]
                store_names = [s.get('store_code', '') for s in stores_sorted]
                connections = [s.get('connection_count', 0) for s in stores_sorted]

                bars = ax3.bar(store_names, connections, color='#2196F3')
                ax3.set_title('Top 5 Tiendas Más Activas')
                ax3.set_ylabel('Conexiones')
                ax3.tick_params(axis='x', rotation=45)

                # Agregar valores
                for bar in bars:
                    height = bar.get_height()
                    ax3.text(bar.get_x() + bar.get_width() / 2., height,
                             f'{int(height)}', ha='center', va='bottom')

            # Gráfico 4: Tendencias temporales (ejemplo)
            ax4 = axes[1, 1]
            # Datos de ejemplo para tendencia
            days = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
            connections_day = [45, 52, 48, 55, 53, 40, 35]

            ax4.plot(days, connections_day, marker='o', color='#9C27B0', linewidth=2)
            ax4.fill_between(days, connections_day, color='#E1BEE7', alpha=0.3)
            ax4.set_title('Conexiones por Día (Ejemplo)')
            ax4.set_ylabel('Conexiones')
            ax4.grid(True, alpha=0.3)

            # Ajustar layout
            plt.tight_layout()

            # Guardar en BytesIO
            buf = io.BytesIO()
            plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
            buf.seek(0)
            plt.close()

            return buf

        except Exception as e:
            logger.error(f"Error creando gráficos: {e}")
            return None


# Handler para el comando /reportes
async def handle_report_command(bot, message):
    """Maneja el comando /reportes"""
    try:
        user_id = message.from_user.id
        username = message.from_user.username or message.from_user.first_name
        chat_id = message.chat.id

        await bot.send_message(
            chat_id=chat_id,
            text="📊 *Generando reporte completo del sistema...*\n\nEsto puede tomar unos momentos.",
            parse_mode="Markdown"
        )

        # Generar reporte
        report_generator = ReportGenerator()
        report_file = report_generator.generate_usage_report(user_id, username)

        if report_file and os.path.exists(report_file):
            # Enviar archivo
            with open(report_file, 'rb') as f:
                await bot.send_document(
                    chat_id=chat_id,
                    document=f,
                    caption=f"📋 *Reporte generado*\n\n✅ Archivo: `{os.path.basename(report_file)}`\n👤 Generado por: {username}\n🕒 Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    parse_mode="Markdown"
                )

            # Registrar en logs
            logger.info(f"Reporte enviado a usuario {user_id} ({username})")

        else:
            await bot.send_message(
                chat_id=chat_id,
                text="❌ *Error generando reporte*\n\nNo se pudo generar el archivo de reporte.",
                parse_mode="Markdown"
            )

    except Exception as e:
        logger.error(f"Error en comando /reportes: {e}")
        await bot.send_message(
            chat_id=message.chat.id,
            text=f"❌ *Error generando reporte:*\n`{str(e)[:200]}`",
            parse_mode="Markdown"
        )